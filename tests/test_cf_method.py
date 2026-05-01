"""Tests para Issue 2: detección del método CF + filtrado condicional en exporters.

SMV expone el método (directo/indirecto) en el campo `MetodoFlujoEfectivo`
de cada fila. Cada empresa publica solo el bloque de cuentas correspondiente
a su método; el otro bloque queda con valores ausentes/cero.
"""
from pathlib import Path

import pytest

from smv_peru import fetch_estados_financieros
from smv_peru.client import _detect_cf_method

FIXTURES = Path(__file__).parent / "fixtures"


# ---------------------------------------------------------------------------
# _detect_cf_method: heurística sobre 'MetodoFlujoEfectivo'
# ---------------------------------------------------------------------------

def test_detect_directo():
    rows = [{"MetodoFlujoEfectivo": "Método Directo"}]
    assert _detect_cf_method(rows) == "directo"


def test_detect_indirecto():
    rows = [{"MetodoFlujoEfectivo": "Método Indirecto"}]
    assert _detect_cf_method(rows) == "indirecto"


def test_detect_robust_to_encoding_issues():
    # SMV a veces devuelve sin tilde ('M todo Directo'); confirmamos robustez
    assert _detect_cf_method([{"MetodoFlujoEfectivo": "M todo Directo"}]) == "directo"
    assert _detect_cf_method([{"MetodoFlujoEfectivo": "M todo Indirecto"}]) == "indirecto"


def test_detect_returns_none_for_empty_or_unknown():
    assert _detect_cf_method([]) is None
    assert _detect_cf_method([{"MetodoFlujoEfectivo": ""}]) is None
    assert _detect_cf_method([{"MetodoFlujoEfectivo": "otra cosa"}]) is None


# ---------------------------------------------------------------------------
# fetch_eeff: cada período expone cf_method
# ---------------------------------------------------------------------------

def test_alicorp_es_metodo_directo():
    # Alicorp publica CF con método directo (es la mayoría de empresas peruanas)
    datos = fetch_estados_financieros("ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES)
    assert datos is not None
    assert datos["periods"][0]["cf_method"] == "directo"


def test_metodo_directo_deja_indirectos_en_none():
    # Alicorp (directo) no debe tener valores en los campos exclusivos del
    # bloque indirecto. NOTA: `ni_before_tax_cf` (3D05ST) sí viene poblado
    # incluso en empresas directas porque SMV lo calcula como subtotal
    # automático — por eso queda fuera de este chequeo.
    datos = fetch_estados_financieros("ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES)
    p = datos["periods"][0]
    for f in ("dna", "fx_adjustment_cf", "ppe_disposal_cf", "other_non_cash_cf",
              "change_in_receivables", "change_in_other_op_assets",
              "change_in_inventory", "change_in_payables", "change_in_other_op_liab"):
        assert p[f] is None, f"{f} debería ser None en empresa con método directo"


def test_3d05st_aparece_aun_en_metodo_directo():
    # SMV publica el subtotal 3D05ST también en empresas con método directo
    # (no es exclusivo del bloque indirecto). Lo exponemos siempre.
    datos = fetch_estados_financieros("ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES)
    p = datos["periods"][0]
    assert p["cf_method"] == "directo"
    assert p["ni_before_tax_cf"] is not None


# ---------------------------------------------------------------------------
# Exporters: filtrado condicional según cf_method
# ---------------------------------------------------------------------------

def test_excel_oculta_indirecto_para_empresa_directa(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from smv_peru import to_excel

    datos = fetch_estados_financieros("ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES)
    path = to_excel(datos, tmp_path / "alicorp.xlsx", ticker="ALICORC1")
    wb = load_workbook(path)
    labels = [str(r[0].value or "") for r in wb["EEFF"].iter_rows(min_col=1, max_col=1)]
    flat = "\n".join(labels)
    # Subgrupo y campos del directo SÍ aparecen
    assert "Operación — Método directo" in flat
    assert "Cash from customers" in flat
    # Subgrupo y campos exclusivos del indirecto NO aparecen
    assert "Operación — Método indirecto" not in flat
    assert "Δ Receivables" not in flat
    assert "FX adjustment (non-cash)" not in flat
    # `ni_before_tax_cf` ahora vive en el subgrupo Indirecto y se oculta para
    # empresas con método directo (aunque SMV lo publique como subtotal). El
    # dato sigue accesible vía `period["ni_before_tax_cf"]`.
    assert "Pretax income (CF starting point)" not in flat
    # interest_paid y taxes_paid ahora viven en el subgrupo Directo
    assert "Interest paid (total)" in flat
    assert "Taxes paid" in flat


def test_csv_oculta_indirecto_para_empresa_directa(tmp_path):
    from smv_peru import to_csv
    datos = fetch_estados_financieros("ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES)
    path = to_csv(datos, tmp_path / "alicorp.csv")
    body = path.read_text()
    assert "Cash from customers" in body
    assert "Δ Receivables" not in body
    assert "-- Operación — Método indirecto --" not in body
