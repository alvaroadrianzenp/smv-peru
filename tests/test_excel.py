"""Tests para to_excel: exporta el output a un archivo Excel.

openpyxl es dep opcional (extras_require=excel). Estos tests dependen de
que openpyxl esté instalado en el entorno de dev (uv sync --extra excel).
"""
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook

from smv_peru import fetch_estados_financieros, to_excel

FIXTURES = Path(__file__).parent / "fixtures"


# ---------------------------------------------------------------------------
# Validaciones de input
# ---------------------------------------------------------------------------

def test_to_excel_raises_on_empty_result(tmp_path):
    with pytest.raises(ValueError, match="periods"):
        to_excel({}, tmp_path / "out.xlsx")


def test_to_excel_raises_on_no_periods(tmp_path):
    with pytest.raises(ValueError, match="periods"):
        to_excel({"periods": []}, tmp_path / "out.xlsx")


# ---------------------------------------------------------------------------
# Esquema 2D (industriales): Alicorp 2021-2023
# ---------------------------------------------------------------------------

def test_to_excel_2d_creates_file(tmp_path):
    datos = fetch_estados_financieros(
        "ALICORC1", desde=2021, hasta=2023, cache_dir=FIXTURES,
    )
    out = to_excel(datos, tmp_path / "alicorp.xlsx", ticker="ALICORC1")
    assert out.exists()
    assert out.stat().st_size > 1000  # archivo razonable, no vacío


def test_to_excel_2d_has_expected_structure(tmp_path):
    datos = fetch_estados_financieros(
        "ALICORC1", desde=2021, hasta=2023, cache_dir=FIXTURES,
    )
    path = to_excel(datos, tmp_path / "alicorp.xlsx", ticker="ALICORC1")
    wb = load_workbook(path)

    assert "EEFF" in wb.sheetnames
    assert "Cuentas adicionales (raw)" not in wb.sheetnames  # off por default

    ws = wb["EEFF"]
    # Header con metadata
    assert "ALICORC1" in str(ws["A1"].value)
    assert "2D" in str(ws["A2"].value)

    # Header de columnas en fila 6: períodos en orden
    assert ws.cell(row=6, column=2).value == "2021"
    assert ws.cell(row=6, column=3).value == "2022"
    assert ws.cell(row=6, column=4).value == "2023"


def test_to_excel_2d_contains_revenue_value(tmp_path):
    datos = fetch_estados_financieros(
        "ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES,
    )
    path = to_excel(datos, tmp_path / "alicorp.xlsx", ticker="ALICORC1")
    wb = load_workbook(path)
    ws = wb["EEFF"]
    # Buscar la fila de Revenue
    revenue_row = None
    for row in ws.iter_rows(min_row=7, values_only=False):
        if row[0].value and "Revenue" in str(row[0].value):
            revenue_row = row
            break
    assert revenue_row is not None
    # El valor en la columna 2 (primer año) debe ser el revenue exacto
    assert revenue_row[1].value == 13_655_764


# ---------------------------------------------------------------------------
# Esquema 2F (bancos): BBVA 2024
# ---------------------------------------------------------------------------

def test_to_excel_2f_uses_banking_sections(tmp_path):
    datos = fetch_estados_financieros(
        "BBVAC1", desde=2024, hasta=2024, cache_dir=FIXTURES,
    )
    path = to_excel(datos, tmp_path / "bbva.xlsx", ticker="BBVAC1")
    wb = load_workbook(path)
    ws = wb["EEFF"]
    # Schema metadata indica 2F
    assert "2F" in str(ws["A2"].value)

    # Debe aparecer "RATIOS BANCARIOS" como sección
    found_banking = False
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] and "RATIOS BANCARIOS" in str(row[0]):
            found_banking = True
            break
    assert found_banking


def test_to_excel_2f_contains_loans_net(tmp_path):
    datos = fetch_estados_financieros(
        "BBVAC1", desde=2024, hasta=2024, cache_dir=FIXTURES,
    )
    path = to_excel(datos, tmp_path / "bbva.xlsx", ticker="BBVAC1")
    wb = load_workbook(path)
    ws = wb["EEFF"]
    found = None
    for row in ws.iter_rows(min_row=7, values_only=False):
        if row[0].value and "Loans, net (total)" in str(row[0].value):
            found = row
            break
    assert found is not None
    # 74,118,352 (PDF auditado BBVA 2024)
    assert found[1].value == 74_118_352


# ---------------------------------------------------------------------------
# include_raw
# ---------------------------------------------------------------------------

def test_to_excel_include_raw_creates_second_sheet(tmp_path):
    datos = fetch_estados_financieros(
        "ALICORC1", desde=2023, hasta=2023, cache_dir=FIXTURES,
    )
    path = to_excel(
        datos, tmp_path / "alicorp_full.xlsx",
        ticker="ALICORC1", include_raw=True,
    )
    wb = load_workbook(path)
    assert "Cuentas adicionales (raw)" in wb.sheetnames

    ws_raw = wb["Cuentas adicionales (raw)"]
    # La hoja debe tener al menos algunas filas de raw_accounts
    assert ws_raw.max_row > 5
    # Header en fila 1
    assert ws_raw.cell(row=1, column=1).value == "Código SMV"


# ---------------------------------------------------------------------------
# Trimestral
# ---------------------------------------------------------------------------

def test_to_excel_trimestral_uses_compact_labels(tmp_path):
    datos = fetch_estados_financieros(
        "ALICORC1", desde=2023, hasta=2023,
        periodicidad="trimestral", cache_dir=FIXTURES,
    )
    path = to_excel(datos, tmp_path / "alicorp_q.xlsx", ticker="ALICORC1")
    wb = load_workbook(path)
    ws = wb["EEFF"]
    # Headers de columnas: 2023Q1, 2023Q2, 2023Q3, 2023Q4
    assert ws.cell(row=6, column=2).value == "2023Q1"
    assert ws.cell(row=6, column=5).value == "2023Q4"
