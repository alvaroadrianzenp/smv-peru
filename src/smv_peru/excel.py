"""excel.py — exportar resultados de fetch_estados_financieros a un archivo Excel.

Genera una plantilla histórica lista para usar en modelos financieros: filas por
campo agrupadas en secciones (P&L, Balance, Cash Flow, Ratios, YoY) y columnas
por período en orden cronológico. Soporta ambos esquemas (2D industriales y 2F
bancos) automáticamente según el campo ``schema`` del output.

**Requiere openpyxl** como dependencia opcional. Instalar con:

    pip install smv-peru[excel]

Ejemplo:

    from smv_peru import fetch_estados_financieros, to_excel
    datos = fetch_estados_financieros("ALICORC1", desde=2019, hasta=2024)
    to_excel(datos, "alicorp_2019_2024.xlsx", ticker="ALICORC1")
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .empresas import EMPRESAS


# ---------------------------------------------------------------------------
# Esquema de filas: lista de secciones, cada una con sus campos.
# Cada item de la lista interna puede ser:
#   * Campo:    (field_key, label, fmt)        — 3-tuple
#   * Subgrupo: (subgroup_label, "subheader")  — 2-tuple, encabezado intermedio
# Tipos: "money" (separador miles), "pct" (porcentaje), "ratio" (Nx),
#        "decimal" (2 decimales sin sufijo)
# ---------------------------------------------------------------------------

# Campos del CF que solo existen para el método directo o el indirecto.
# Se ocultan en el Excel cuando la empresa usa el método contrario.
_DIRECT_CF_FIELDS = frozenset({
    "cash_from_customers", "interest_received_op",
    "cash_to_suppliers", "cash_to_employees",
    "interest_paid", "taxes_paid",
})
_INDIRECT_CF_FIELDS = frozenset({
    "ni_before_tax_cf", "dna",
    "fx_adjustment_cf", "ppe_disposal_cf", "other_non_cash_cf",
    "change_in_receivables", "change_in_other_op_assets",
    "change_in_inventory", "change_in_payables", "change_in_other_op_liab",
})

SECTIONS_2D: list[tuple[str, list[tuple[str, str, str]]]] = [
    ("ESTADO DE RESULTADOS", [
        ("revenue", "Revenue", "money"),
        ("cogs", "Cost of goods sold", "money"),
        ("gross_profit", "Gross profit", "money"),
        ("gross_margin", "Gross margin", "pct"),
        ("selling_expenses", "Selling expenses", "money"),
        ("admin_expenses", "Admin expenses", "money"),
        ("other_op_income", "Other operating income", "money"),
        ("other_op_expenses", "Other operating expenses", "money"),
        ("operating_income", "Operating income (EBIT)", "money"),
        ("operating_margin", "Operating margin", "pct"),
        ("ebitda", "EBITDA (real, requires D&A)", "money"),
        ("interest_income", "Interest income", "money"),
        ("interest_expense", "Interest expense", "money"),
        ("fx_gain_loss", "FX gain/loss (net)", "money"),
        ("pretax_income", "Pretax income", "money"),
        ("income_tax", "Income tax", "money"),
        ("effective_tax_rate", "Effective tax rate", "pct"),
        ("net_income", "Net income", "money"),
        ("net_income_to_parent", "  Attributable to parent", "money"),
        ("minority_interest", "  Minority interest", "money"),
        ("net_margin", "Net margin", "pct"),
        ("eps", "EPS (basic)", "decimal"),
        ("eps_diluted", "EPS (diluted)", "decimal"),
    ]),
    ("BALANCE — Activos", [
        ("cash", "Cash & equivalents", "money"),
        ("accounts_receivable", "Accounts receivable", "money"),
        ("inventory", "Inventory", "money"),
        ("biological_assets", "Biological assets", "money"),
        ("current_assets", "Current assets (subtotal)", "money"),
        ("ppe", "PP&E", "money"),
        ("intangibles", "Intangibles", "money"),
        ("goodwill", "Goodwill", "money"),
        ("investment_property", "Investment property", "money"),
        ("equity_investments", "Equity-method investments", "money"),
        ("noncurrent_assets", "Non-current assets (subtotal)", "money"),
        ("total_assets", "Total assets", "money"),
    ]),
    ("BALANCE — Pasivos y Patrimonio", [
        ("accounts_payable", "Accounts payable", "money"),
        ("debt_short_term", "Short-term debt", "money"),
        ("employee_benefits", "Employee benefits provision", "money"),
        ("current_liab", "Current liabilities (subtotal)", "money"),
        ("debt_long_term", "Long-term debt", "money"),
        ("noncurrent_liab", "Non-current liabilities (subtotal)", "money"),
        ("total_liabilities", "Total liabilities", "money"),
        ("share_capital", "Share capital", "money"),
        ("investment_shares", "Investment shares (Perú-specific)", "money"),
        ("reserves", "Reserves", "money"),
        ("retained_earnings", "Retained earnings", "money"),
        ("equity_to_parent", "  Attributable to parent", "money"),
        ("minority_equity", "  Minority interest", "money"),
        ("equity", "Total equity", "money"),
    ]),
    ("CASH FLOW", [
        ("Operación — Método directo", "subheader"),
        ("cash_from_customers", "Cash from customers", "money"),
        ("interest_received_op", "Interest received (op)", "money"),
        ("cash_to_suppliers", "Cash to suppliers", "money"),
        ("cash_to_employees", "Cash to employees", "money"),
        ("interest_paid", "Interest paid (total)", "money"),
        ("taxes_paid", "Taxes paid", "money"),
        ("Operación — Método indirecto", "subheader"),
        ("ni_before_tax_cf", "Pretax income (CF starting point)", "money"),
        ("dna", "D&A", "money"),
        ("fx_adjustment_cf", "FX adjustment (non-cash)", "money"),
        ("ppe_disposal_cf", "PP&E disposal (gain/loss, non-cash)", "money"),
        ("other_non_cash_cf", "Other non-cash items", "money"),
        ("change_in_receivables", "Δ Receivables", "money"),
        ("change_in_other_op_assets", "Δ Other operating assets", "money"),
        ("change_in_inventory", "Δ Inventory", "money"),
        ("change_in_payables", "Δ Payables", "money"),
        ("change_in_other_op_liab", "Δ Other operating liabilities", "money"),
        ("Total operación", "subheader"),
        ("operating_cf", "Operating cash flow (subtotal)", "money"),
        ("Inversión", "subheader"),
        ("ppe_proceeds", "PP&E proceeds", "money"),
        ("capex_ppe", "Capex PP&E", "money"),
        ("capex_intangibles", "Capex intangibles", "money"),
        ("capex_total", "Capex total", "money"),
        ("dividends_received", "Dividends received (inv)", "money"),
        ("investing_cf", "Investing cash flow (subtotal)", "money"),
        ("Financiamiento", "subheader"),
        ("debt_issued", "Debt issued", "money"),
        ("debt_repaid", "Debt repaid", "money"),
        ("equity_issued", "Equity issued", "money"),
        ("dividends_paid", "Dividends paid", "money"),
        ("financing_cf", "Financing cash flow (subtotal)", "money"),
        ("Resultado", "subheader"),
        ("fcf", "Free cash flow", "money"),
        ("cash_change_pre_fx", "Net change in cash (pre-FX)", "money"),
        ("fx_effect_cash", "FX effect on cash", "money"),
        ("net_change_in_cash", "Net change in cash (total)", "money"),
        ("start_cash", "Start-of-period cash", "money"),
        ("end_cash", "End-of-period cash", "money"),
    ]),
    ("EBITDA Y MÉTRICAS DE CRÉDITO", [
        ("ebitda", "EBITDA", "money"),
        ("ebitda_margin", "EBITDA margin", "pct"),
        ("total_debt", "Total debt", "money"),
        ("net_debt", "Net debt", "money"),
        ("debt_to_ebitda", "Debt / EBITDA", "ratio"),
        ("net_debt_to_ebitda", "Net debt / EBITDA", "ratio"),
        ("interest_coverage_ebitda", "EBITDA / Interest expense", "ratio"),
    ]),
    ("RATIOS Y MÉTRICAS DERIVADAS", [
        ("current_ratio", "Current ratio", "ratio"),
        ("quick_ratio", "Quick ratio", "ratio"),
        ("interest_coverage", "Interest coverage (EBIT/int.expense)", "ratio"),
        ("capex_intensity", "Capex intensity (capex/revenue)", "pct"),
        ("payout_ratio", "Payout ratio", "pct"),
        ("roe", "ROE (avg equity)", "pct"),
        ("roic", "ROIC (avg invested capital)", "pct"),
    ]),
    ("CRECIMIENTO YoY", [
        ("revenue_yoy", "Revenue growth YoY", "pct"),
        ("net_income_yoy", "Net income growth YoY", "pct"),
        ("equity_yoy", "Equity growth YoY", "pct"),
    ]),
]


SECTIONS_2F: list[tuple[str, list[tuple[str, str, str]]]] = [
    ("ESTADO DE RESULTADOS (Banking P&L)", [
        ("Margen financiero", "subheader"),
        ("interest_income", "Interest income", "money"),
        ("interest_expense", "Interest expense", "money"),
        ("nii_pure", "Net interest income (pure)", "money"),
        ("net_interest_income", "Margen Bruto SMV (oficial)", "money"),
        ("loan_loss_provisions", "Loan loss provisions", "money"),
        ("nii_after_provisions", "Margen Financiero Neto (post-LLP)", "money"),
        ("Servicios y trading", "subheader"),
        ("fee_income_net", "Fee income (net)", "money"),
        ("op_revenue_after_fees", "Margen post-servicios (SBS only)", "money"),
        ("trading_income", "Trading income (ROF)", "money"),
        ("op_revenue_total", "Margen Operacional (incl. ROF, SBS only)", "money"),
        ("Resultado operativo", "subheader"),
        ("operating_expenses", "Operating expenses", "money"),
        ("op_income_pre_op_provisions", "Margen Op. Neto (pre otras prov., SBS only)", "money"),
        ("operating_income", "Operating income (Resultado de Operación)", "money"),
        ("non_op_items", "Non-operating items", "money"),
        ("Resultado del ejercicio", "subheader"),
        ("pretax_income", "Pretax income", "money"),
        ("income_tax", "Income tax", "money"),
        ("effective_tax_rate", "Effective tax rate", "pct"),
        ("net_income", "Net income", "money"),
        ("eps", "EPS (basic)", "decimal"),
        ("eps_diluted", "EPS (diluted)", "decimal"),
    ]),
    ("BALANCE — Activos", [
        ("cash", "Cash (Disponibles)", "money"),
        ("interbank_funds", "Interbank funds (asset)", "money"),
        ("investments_fvtpl", "Investments at FVTPL", "money"),
        ("investments_afs", "Investments AFS", "money"),
        ("investments_htm", "Investments HTM", "money"),
        ("loans_st", "Loans, net (current)", "money"),
        ("loans_lt", "Loans, net (non-current)", "money"),
        ("loans_net", "Loans, net (total)", "money"),
        ("performing_loans", "Performing loans (gross)", "money"),
        ("refinanced_loans", "Refinanced loans", "money"),
        ("overdue_loans", "Overdue loans", "money"),
        ("judicial_loans", "Judicial loans", "money"),
        ("gross_loans", "Gross loans (sum of components)", "money"),
        ("repossessed_assets", "Repossessed assets (foreclosed)", "money"),
        ("ppe", "PP&E", "money"),
        ("intangibles", "Intangibles", "money"),
        ("total_assets", "Total assets", "money"),
    ]),
    ("BALANCE — Pasivos y Patrimonio", [
        ("deposits", "Deposits (Obligaciones con público)", "money"),
        ("interbank_funds_payable", "Interbank funds (liability)", "money"),
        ("deposits_financial_system", "Deposits from financial system", "money"),
        ("financial_debt_st", "Financial debt (current)", "money"),
        ("financial_debt_lt", "Financial debt (non-current)", "money"),
        ("total_liabilities", "Total liabilities", "money"),
        ("share_capital", "Share capital", "money"),
        ("reserves", "Reserves", "money"),
        ("retained_earnings", "Retained earnings", "money"),
        ("equity", "Total equity", "money"),
    ]),
    ("CASH FLOW", [
        ("Operación", "subheader"),
        ("dna", "Depreciation & amortization", "money"),
        ("deposits_change", "Δ Deposits", "money"),
        ("loans_change", "Δ Loans", "money"),
        ("operating_cf", "Operating cash flow (subtotal)", "money"),
        ("Inversión", "subheader"),
        ("investing_cf", "Investing cash flow (subtotal)", "money"),
        ("Financiamiento", "subheader"),
        ("dividends_paid", "Dividends paid (absolute)", "money"),
        ("financing_cf", "Financing cash flow (subtotal)", "money"),
        ("Resultado", "subheader"),
        ("cash_change_pre_fx", "Net change in cash (pre-FX)", "money"),
        ("fx_effect_cash", "FX effect on cash", "money"),
        ("net_change_in_cash", "Net change in cash (total)", "money"),
        ("start_cash", "Start-of-period cash", "money"),
        ("end_cash", "End-of-period cash", "money"),
    ]),
    ("RATIOS BANCARIOS", [
        ("nim", "NIM (vs avg loans)", "pct"),
        ("efficiency_ratio", "Efficiency ratio", "pct"),
        ("npl_ratio", "NPL ratio (proxy)", "pct"),
        ("loan_to_deposit_ratio", "Loan-to-deposit ratio", "pct"),
        ("cost_of_risk", "Cost of risk", "pct"),
        ("equity_to_assets", "Equity / Total assets", "pct"),
        ("roa", "ROA (avg total assets)", "pct"),
        ("roe", "ROE (avg equity)", "pct"),
    ]),
    ("CRECIMIENTO YoY", [
        ("interest_income_yoy", "Interest income growth YoY", "pct"),
        ("net_income_yoy", "Net income growth YoY", "pct"),
        ("loans_yoy", "Loans growth YoY", "pct"),
        ("deposits_yoy", "Deposits growth YoY", "pct"),
        ("equity_yoy", "Equity growth YoY", "pct"),
    ]),
]


def _period_label(p: dict) -> str:
    """Devuelve etiqueta de período: '2023' o '2023Q3'."""
    if p.get("quarter") is None:
        return str(p["fiscal_year"])
    return f"{p['fiscal_year']}Q{p['quarter']}"


def _format_for(fmt: str) -> str | None:
    """Devuelve el number_format de Excel para un tipo de campo."""
    if fmt == "money":
        return '#,##0;(#,##0);"—"'
    if fmt == "pct":
        return '0.0%;-0.0%;"—"'
    if fmt == "ratio":
        return '0.00"x"'
    if fmt == "decimal":
        return '0.00'
    return None


def _populate_eeff_sheet(ws, periods: list[dict], ticker: str | None) -> None:
    """Puebla una hoja con el layout de EEFF: header + secciones + campos.

    Las cuentas amigables del CF se muestran condicionalmente según el método
    que la empresa usa: si todos los períodos reportan método directo, las
    cuentas del bloque indirecto se omiten (y viceversa). Si hay mixtura
    (raro) o método desconocido, se muestran ambos bloques.
    """
    schema = periods[0].get("schema", "2D")
    sections = SECTIONS_2D if schema == "2D" else SECTIONS_2F

    methods = {p.get("cf_method") for p in periods if p.get("cf_method")}
    show_direct = (not methods) or ("directo" in methods)
    show_indirect = (not methods) or ("indirecto" in methods)

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    subheader_font = Font(bold=True, italic=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    section_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    subheader_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    right_align = Alignment(horizontal="right")

    nombre_emp = ""
    if ticker and ticker in EMPRESAS:
        nombre_emp = EMPRESAS[ticker]["nombre"]
    elif ticker:
        nombre_emp = ticker

    titulo = nombre_emp + (f" ({ticker})" if ticker and nombre_emp != ticker else "")
    ws["A1"] = titulo or "Estados Financieros"
    ws["A1"].font = title_font
    currency = periods[0].get("currency") or "—"
    currency_full = {"PEN": "Soles peruanos (PEN)", "USD": "Dólares (USD)"}.get(currency, currency)
    ws["A2"] = f"Moneda: {currency_full}"
    ws["A3"] = f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
    ws["A4"] = "Montos en miles. Ratios como porcentajes."

    HEADER_ROW = 6
    LABEL_COL = 1
    FIRST_DATA_COL = 2

    period_labels = [_period_label(p) for p in periods]
    label_cell = ws.cell(row=HEADER_ROW, column=LABEL_COL, value="")
    label_cell.fill = header_fill
    for i, label in enumerate(period_labels):
        cell = ws.cell(row=HEADER_ROW, column=FIRST_DATA_COL + i, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = right_align

    current_row = HEADER_ROW + 1
    for section_name, items in sections:
        c = ws.cell(row=current_row, column=LABEL_COL, value=section_name)
        c.font = bold
        c.fill = section_fill
        for col in range(FIRST_DATA_COL, FIRST_DATA_COL + len(periods)):
            ws.cell(row=current_row, column=col).fill = section_fill
        current_row += 1

        for item in items:
            # Subgrupo (label + "subheader"): encabezado intermedio en cursiva.
            if len(item) == 2 and item[1] == "subheader":
                subheader_label = item[0]
                # Ocultar el subgrupo "Operación — Método directo/indirecto"
                # si la empresa no usa ese método.
                if subheader_label.startswith("Operación — Método directo") and not show_direct:
                    continue
                if subheader_label.startswith("Operación — Método indirecto") and not show_indirect:
                    continue
                cell = ws.cell(row=current_row, column=LABEL_COL, value="  " + subheader_label)
                cell.font = subheader_font
                cell.fill = subheader_fill
                for col in range(FIRST_DATA_COL, FIRST_DATA_COL + len(periods)):
                    ws.cell(row=current_row, column=col).fill = subheader_fill
                current_row += 1
                continue

            # Campo (3-tuple).
            field_key, label, fmt = item
            if field_key in _DIRECT_CF_FIELDS and not show_direct:
                continue
            if field_key in _INDIRECT_CF_FIELDS and not show_indirect:
                continue
            ws.cell(row=current_row, column=LABEL_COL, value="    " + label)
            num_fmt = _format_for(fmt)
            for i, p in enumerate(periods):
                value = p.get(field_key)
                cell = ws.cell(row=current_row, column=FIRST_DATA_COL + i)
                if value is None:
                    cell.value = "—"
                    cell.alignment = right_align
                else:
                    cell.value = value
                    if num_fmt:
                        cell.number_format = num_fmt
                    cell.alignment = right_align
            current_row += 1

        current_row += 1

    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 38
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 14

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=FIRST_DATA_COL)


def _populate_raw_sheet(ws, periods: list[dict]) -> None:
    """Puebla una hoja con raw_accounts (códigos SMV + descripción + montos)."""
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    right_align = Alignment(horizontal="right")

    period_labels = [_period_label(p) for p in periods]
    all_codes: dict[str, str] = {}
    for p in periods:
        for code, info in p.get("raw_accounts", {}).items():
            if code not in all_codes:
                all_codes[code] = info.get("nombre", code) or code

    ws.cell(row=1, column=1, value="Código SMV").font = bold
    ws.cell(row=1, column=2, value="Descripción oficial").font = bold
    for i, label in enumerate(period_labels):
        cell = ws.cell(row=1, column=3 + i, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = right_align
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2).fill = header_fill

    for r, (code, nombre) in enumerate(sorted(all_codes.items()), start=2):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=nombre)
        for i, p in enumerate(periods):
            if code in p.get("raw_accounts", {}):
                cell = ws.cell(row=r, column=3 + i, value=p["raw_accounts"][code]["monto"])
                cell.number_format = '#,##0;(#,##0);"—"'
                cell.alignment = right_align

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.freeze_panes = ws.cell(row=2, column=3)


def to_excel(
    result: dict,
    filepath: str | Path,
    include_raw: bool = False,
    ticker: str | None = None,
) -> Path:
    """Exporta el resultado de ``fetch_estados_financieros`` o ``fetch_multi`` a un archivo Excel.

    Detecta automáticamente si ``result`` es de una empresa (output de
    ``fetch_estados_financieros``) o múltiples (output de ``fetch_multi``).

    **Single empresa:** una hoja "EEFF" con secciones P&L/Balance/CF/Ratios/YoY.
    **Multi-empresa:** una hoja por ticker, nombrada con el ticker.

    Layout: filas = campos amigables agrupados por sección. Columnas = períodos
    cronológicos. Header con metadata (ticker, esquema, fecha).

    Args:
        result: dict de ``fetch_estados_financieros`` (con ``periods``) o
            ``fetch_multi`` (dict de dicts).
        filepath: ruta del archivo Excel a generar (.xlsx).
        include_raw: si True, agrega hoja ``Cuentas adicionales`` con raw_accounts.
            Para multi-empresa, una hoja raw por ticker con sufijo ``_raw``.
        ticker: ticker BVL (opcional, solo single-empresa). Enriquece header.

    Returns:
        Path absoluto del archivo generado.

    Raises:
        ImportError: si openpyxl no está instalado. ``pip install smv-peru[excel]``.
        ValueError: si ``result`` está vacío.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl no está instalado. Para exportar a Excel ejecuta: "
            "pip install smv-peru[excel]"
        )
    if not result:
        raise ValueError("'result' está vacío")

    # Detectar shape: single (con 'periods') vs multi (dict de dicts)
    is_multi = "periods" not in result

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja default vacía

    if is_multi:
        # Multi-empresa: filtrar tickers válidos (con datos)
        valid_items = [(t, r) for t, r in result.items() if r and r.get("periods")]
        if not valid_items:
            raise ValueError(
                "'result' multi-empresa no tiene ningún ticker con datos"
            )
        for ticker_i, single_result in valid_items:
            ws = wb.create_sheet(ticker_i[:31])  # Excel limit: 31 chars sheet name
            _populate_eeff_sheet(ws, single_result["periods"], ticker_i)
            if include_raw:
                ws_raw = wb.create_sheet(f"{ticker_i[:25]}_raw")
                _populate_raw_sheet(ws_raw, single_result["periods"])
    else:
        if not result.get("periods"):
            raise ValueError("'result' no contiene 'periods'")
        ws = wb.create_sheet("EEFF")
        _populate_eeff_sheet(ws, result["periods"], ticker)
        if include_raw:
            ws_raw = wb.create_sheet("Cuentas adicionales (raw)")
            _populate_raw_sheet(ws_raw, result["periods"])

    filepath = Path(filepath)
    wb.save(filepath)
    return filepath
